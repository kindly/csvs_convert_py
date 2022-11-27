import unittest
import tempfile
import glob
from openpyxl import load_workbook
import pandas
import sqlite3

import csvs_convert


class TestXLSX(unittest.TestCase):

    def test_large(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.datapackage_to_xlsx(f'{tmpdir}/output_large.xlsx', 'fixtures/large')
            wb = load_workbook(filename = f'{tmpdir}/output_large.xlsx')
            titles = [a.title for a in wb.worksheets]
            self.assertEqual(titles, ['daily_16', 'data', 'looooooo_data_weather_name'])

    def test_titles(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.datapackage_to_xlsx(f'{tmpdir}/output_large.xlsx', 'fixtures/large', use_titles=True)
            wb = load_workbook(filename = f'{tmpdir}/output_large.xlsx')
            titles = [a.title for a in wb.worksheets]
            self.assertEqual(titles, ['daily_16', 'data', 'Loooooo_data_weather_Title'])

    def test_from_csvs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.csvs_to_xlsx(f'{tmpdir}/output_from_csvs.xlsx', ["fixtures/large/csv/data.csv"])
            wb = load_workbook(filename = f'{tmpdir}/output_from_csvs.xlsx')
            titles = [a.title for a in wb.worksheets]
            self.assertEqual(titles, ['data'])

class TestParquet(unittest.TestCase):

    def test_large(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.datapackage_to_parquet(f'{tmpdir}/output', 'fixtures/large')
            files = glob.glob(f'{tmpdir}/output/*')
            self.assertEqual(sorted([f.split('/')[-1] for f in files]),
                sorted(['looooooooooooong_data_weather_name.parquet',
                        'data.parquet',
                        'daily_16.parquet']))

            lengths = []
            for parquet_file in files:
                frame = pandas.read_parquet(parquet_file)
                lengths.append(len(frame))
            self.assertEqual(sorted(lengths), [101, 1679, 1679])

    def test_from_csvs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.csvs_to_parquet(f'{tmpdir}/output', ["fixtures/large/csv/data.csv"])

            files = glob.glob(f'{tmpdir}/output/*')
            self.assertEqual(sorted([f.split('/')[-1] for f in files]),
                sorted(['data.parquet']))


class TestSqlite(unittest.TestCase):
    def test_large(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.datapackage_to_sqlite(f'{tmpdir}/output.db', 'fixtures/large')
            con = sqlite3.connect(f'{tmpdir}/output.db')
            cur = con.cursor()

            tables = []
            for row in cur.execute("SELECT name FROM sqlite_master WHERE type='table';"):
                tables.append(row[0])

            self.assertEqual(sorted(tables),
                sorted(['Looooooooooooong_data_weather_Title',
                        'daily_16',
                        'data']))

    def test_from_csvs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.csvs_to_sqlite(f'{tmpdir}/output.db', ["fixtures/large/csv/data.csv"])

            con = sqlite3.connect(f'{tmpdir}/output.db')
            cur = con.cursor()
            tables = []
            for row in cur.execute("SELECT name FROM sqlite_master WHERE type='table';"):
                tables.append(row[0])

            self.assertEqual(sorted(tables), ['data'])


class TestMerge(unittest.TestCase):
    def test_merge(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.merge_datapackage(f'{tmpdir}', ['fixtures/base_datapackage', 'fixtures/add_field'])
            with open(f'{tmpdir}/csv/games.csv') as f:
                self.assertEqual(f.readlines(),["id,title,title2\n",
                                                "1,game1_base,\n",
                                                "2,game2_base,\n",
                                                "2,,game1_add_field\n"])


class TestPostgres(unittest.TestCase):
    def test_large(self):
        csvs_convert.datapackage_to_postgres(f'postgres://test@localhost/test', 'fixtures/large', drop=True, schema='test1')

    def test_from_csvs(self):
        csvs_convert.csvs_to_postgres(f'postgres://test@localhost/test', ["fixtures/large/csv/data.csv"], drop=True, schema='test2')


class TestPostgres(unittest.TestCase):
    def test_create_datapackage(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            csvs_convert.csvs_to_datapackage(f'{tmpdir}/datapackage.json', ["fixtures/large/csv/data.csv"])
            with open(f'{tmpdir}/datapackage.json') as f:
                f.read()
